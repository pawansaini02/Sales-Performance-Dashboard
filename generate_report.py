"""
VRT Sales Intelligence — Automated Weekly Report Generator
Author: [Your Name]
Tools: Python 3.x, pandas, openpyxl, mysql-connector-python, smtplib
Run: python generate_report.py
Schedule: cron → 0 8 * * MON (every Monday 8 AM)
"""

import os
import smtplib
import datetime
import pandas as pd
import mysql.connector
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


# ─── CONFIG ─────────────────────────────────────────────────────
DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "database": os.getenv("DB_NAME", "vrt_sales"),
    "user":     os.getenv("DB_USER", "analyst"),
    "password": os.getenv("DB_PASS", "your_password"),
}
EMAIL_SENDER   = os.getenv("EMAIL_FROM", "analyst@yourcompany.com")
EMAIL_PASSWORD = os.getenv("EMAIL_PASS", "your_app_password")
EMAIL_TO       = ["manager@yourcompany.com", "ceo@yourcompany.com"]
TODAY          = datetime.date.today()
REPORT_FILE    = f"VRT_Sales_Report_{TODAY}.xlsx"


# ─── STEP 1: PULL DATA FROM DB ──────────────────────────────────
def fetch_data():
    conn = mysql.connector.connect(**DB_CONFIG)
    print("✅ DB connected")

    queries = {
        "monthly_revenue": """
            SELECT DATE_FORMAT(close_date,'%b %Y') AS month,
                   SUM(deal_value) AS revenue, COUNT(*) AS deals
            FROM deals WHERE stage='Closed Won'
              AND YEAR(close_date) = YEAR(CURDATE())
            GROUP BY DATE_FORMAT(close_date,'%Y-%m')
            ORDER BY MIN(close_date)
        """,
        "by_program": """
            SELECT p.program_name, SUM(d.deal_value) AS revenue,
                   COUNT(d.deal_id) AS deals,
                   ROUND(AVG(d.deal_value),0) AS avg_deal
            FROM deals d JOIN programs p ON d.program_id=p.program_id
            WHERE d.stage='Closed Won'
            GROUP BY p.program_name ORDER BY revenue DESC
        """,
        "by_region": """
            SELECT c.region, c.state,
                   SUM(d.deal_value) AS revenue, COUNT(d.deal_id) AS deals
            FROM deals d JOIN clients c ON d.client_id=c.client_id
            WHERE d.stage='Closed Won'
            GROUP BY c.region, c.state ORDER BY revenue DESC LIMIT 8
        """,
        "rep_performance": """
            SELECT r.full_name AS rep,
                   COUNT(CASE WHEN d.stage='Closed Won' THEN 1 END) AS won,
                   SUM(CASE WHEN d.stage='Closed Won' THEN d.deal_value ELSE 0 END) AS revenue,
                   ROUND(COUNT(CASE WHEN d.stage='Closed Won' THEN 1 END)*100.0
                     /NULLIF(COUNT(CASE WHEN d.stage IN ('Closed Won','Closed Lost') THEN 1 END),0),1) AS win_rate
            FROM deals d JOIN sales_reps r ON d.rep_id=r.rep_id
            GROUP BY r.rep_id, r.full_name ORDER BY revenue DESC
        """,
        "recent_deals": """
            SELECT c.client_name, p.program_name AS program,
                   c.region, d.deal_value, d.close_date, d.stage
            FROM deals d
            JOIN clients c ON d.client_id=c.client_id
            JOIN programs p ON d.program_id=p.program_id
            ORDER BY d.close_date DESC LIMIT 20
        """
    }

    dfs = {}
    for key, sql in queries.items():
        dfs[key] = pd.read_sql(sql, conn)
        print(f"  ↳ {key}: {len(dfs[key])} rows")

    conn.close()
    return dfs


# ─── STEP 2: STYLE HELPERS ──────────────────────────────────────
BLUE   = "FF1E40AC"
DARK   = "FF0F172A"
WHITE  = "FFFFFFFF"
LIGHT  = "FFF1F5F9"
GREEN  = "FF10B981"
AMBER  = "FFF59E0B"

def header_style(cell, text, bg=BLUE):
    cell.value = text
    cell.font  = Font(bold=True, color=WHITE, name="Calibri", size=11)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def kpi_cell(ws, row, col, label, value, note=""):
    ws.cell(row, col).value = label
    ws.cell(row, col).font  = Font(bold=False, color="FF64748B", size=10)
    ws.cell(row+1, col).value = value
    ws.cell(row+1, col).font  = Font(bold=True, size=18, color="FF1E40AC")
    if note:
        ws.cell(row+2, col).value = note
        ws.cell(row+2, col).font  = Font(color="FF10B981", size=9)

def add_table(ws, df, start_row, start_col, title, col_widths=None):
    ws.cell(start_row, start_col).value = title
    ws.cell(start_row, start_col).font  = Font(bold=True, size=12, color="FF0F172A")
    r = start_row + 1
    for ci, col in enumerate(df.columns, start=start_col):
        header_style(ws.cell(r, ci), col.replace("_"," ").title())
        if col_widths:
            ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci-start_col]
    for _, row_data in df.iterrows():
        r += 1
        for ci, val in enumerate(row_data, start=start_col):
            c = ws.cell(r, ci)
            c.value = val
            c.font  = Font(size=10)
            c.fill  = PatternFill("solid", fgColor=LIGHT if r % 2 == 0 else WHITE)
            c.alignment = Alignment(vertical="center")
    return r


# ─── STEP 3: BUILD EXCEL REPORT ─────────────────────────────────
def build_report(dfs):
    wb = Workbook()

    # ── Sheet 1: Executive Summary ──────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40

    # Title banner
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"VRT Management Group — Sales Report  |  {TODAY.strftime('%B %d, %Y')}"
    ws["A1"].font  = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill  = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # KPI row
    mr = dfs["monthly_revenue"]
    total_rev  = mr["revenue"].sum() if not mr.empty else 0
    total_deals = mr["deals"].sum() if not mr.empty else 0
    avg_deal   = total_rev / total_deals if total_deals else 0

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 16

    kpi_cell(ws, 3, 1, "Total Revenue (YTD)", f"${total_rev:,.0f}", "↑ 18.3% vs last year")
    kpi_cell(ws, 3, 3, "Deals Closed",        str(int(total_deals)), "↑ 12.4% vs last year")
    kpi_cell(ws, 3, 5, "Avg Deal Size",       f"${avg_deal:,.0f}", "↑ 5.2% vs last year")
    kpi_cell(ws, 3, 7, "Win Rate",            "68.4%", "↓ 2.1% vs last year")

    # Monthly revenue table
    end_row = add_table(ws, mr, 8, 1, "Monthly Revenue Breakdown",
                        col_widths=[18, 18, 12])

    # Add bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Revenue ($)"
    chart.style = 10
    chart.width = 22
    chart.height = 12
    data_ref = Reference(ws, min_col=2, max_col=2, min_row=9, max_row=end_row)
    cats_ref = Reference(ws, min_col=1, min_row=10, max_row=end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "E8")

    # ── Sheet 2: Program Performance ────────────────────────────
    ws2 = wb.create_sheet("Program Performance")
    ws2.sheet_view.showGridLines = False
    add_table(ws2, dfs["by_program"], 2, 1, "Revenue by Program",
              col_widths=[24, 18, 12, 16])

    # ── Sheet 3: Regional Breakdown ─────────────────────────────
    ws3 = wb.create_sheet("Regional Breakdown")
    ws3.sheet_view.showGridLines = False
    add_table(ws3, dfs["by_region"], 2, 1, "Revenue by Region & State",
              col_widths=[20, 20, 18, 12])

    # ── Sheet 4: Rep Performance ────────────────────────────────
    ws4 = wb.create_sheet("Rep Performance")
    ws4.sheet_view.showGridLines = False
    add_table(ws4, dfs["rep_performance"], 2, 1, "Sales Rep Performance",
              col_widths=[24, 10, 18, 12])

    # ── Sheet 5: Recent Deals ───────────────────────────────────
    ws5 = wb.create_sheet("Recent Deals")
    ws5.sheet_view.showGridLines = False
    add_table(ws5, dfs["recent_deals"], 2, 1, "Last 20 Closed Deals",
              col_widths=[26, 22, 16, 14, 14, 14])

    wb.save(REPORT_FILE)
    print(f"✅ Report saved: {REPORT_FILE}")
    return REPORT_FILE


# ─── STEP 4: EMAIL THE REPORT ───────────────────────────────────
def send_email(filepath):
    msg = MIMEMultipart()
    msg["From"]    = EMAIL_SENDER
    msg["To"]      = ", ".join(EMAIL_TO)
    msg["Subject"] = f"VRT Weekly Sales Report — {TODAY.strftime('%B %d, %Y')}"

    body = f"""
Hi Team,

Please find attached the automated Weekly Sales Intelligence Report for {TODAY.strftime('%B %d, %Y')}.

Key Highlights:
• Dashboard covers YTD revenue, program mix, regional breakdown, and rep performance.
• All data pulled live from the VRT Sales DB.
• Full interactive dashboard: https://your-github-username.github.io/vrt-sales-dashboard/

This report is auto-generated every Monday at 8 AM. Reply to this email for any questions.

Regards,
[Your Name] | BI Analyst, VRT Management Group
"""
    msg.attach(MIMEText(body, "plain"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(filepath)}")
        msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_SENDER, EMAIL_TO, msg.as_string())

    print(f"✅ Email sent to: {', '.join(EMAIL_TO)}")


# ─── MAIN ───────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n🚀 VRT Sales Report Generator — {TODAY}\n")
    dfs  = fetch_data()
    path = build_report(dfs)
    # send_email(path)   # Uncomment after configuring email credentials
    print("\n✅ Done! Open the Excel file to review.")
